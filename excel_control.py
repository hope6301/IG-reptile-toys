import openpyxl
from openpyxl.styles import Font
import os


def download_file(username,usernametext):

    # 獲取使用者的主目錄
    user_home = os.path.expanduser("~")

    # 構建 Downloads 資料夾的路徑
    downloads_path = os.path.join(user_home, "Downloads/ig_file")

    # 確保 指定路徑or資料夾存在
    if not os.path.exists(downloads_path):
        # 如果沒有資料夾，就創建一個資料夾
        os.makedirs(downloads_path)

    # os.chdir 是 python 切換到指定路徑的方法
    os.chdir(downloads_path)

    # 利用 Workbook 建立一個新的工作簿
    workbook = openpyxl.Workbook()

    # 取得第一個工作表
    sheet = workbook.worksheets[0]

    for rowNum in range(len(username)):
        sheet.cell(rowNum+1,1).value = username[rowNum]
        sheet.cell(rowNum+1,2).value = usernametext[rowNum]

    # 儲存檔案
    workbook.save('text.xlsx')

    return("下載完成")











